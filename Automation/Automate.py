from openpyxl import load_workbook
import xlwt
import xlrd

def listOfData(sourceFile):
    workbook = xlrd.open_workbook(sourceFile + '.xlsx')
    sheet = workbook.sheet_by_index(0)
    rows = []

    for i in range(sheet.nrows):
        data = [sheet.cell_value(i, col) for col in range(sheet.ncols)]
        rows.append(data)

    return rows

def copyRow(wbInstance, wsInstance, rowStart, values, fileName):
    print values

    for i in range(7):
        wsInstance.cell(row=rowStart, column=i+1).value = values[i]
    # wbInstance.save("books.xlsx")

def copyTable(sourceFile, destinationFile):

    rows = listOfData(sourceFile)
    wb2 = load_workbook(destinationFile + '.xlsx')
    ws = wb2['Assignment Sheet']

    for i in range(1, 46):
        copyRow(wb2, ws, i+1, rows[i], "books")

    return wb2



print listOfData("excel")

wb2 = load_workbook('SAStats.xlsx')
ws = wb2['Assignment Sheet']

x = copyTable("excel", "books")
x.save("books.xlsx")

# copyRow(2, rows[1], "SAStats")
#
# for i in range(1, 46):
#     copyRow(wb2, ws, i+1, rows[i], "books")
